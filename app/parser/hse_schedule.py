from __future__ import annotations

import html
import logging
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

import httpx
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import xlrd

TIMETABLE_URL = "https://perm.hse.ru/students/timetable/"
FILE_KEYWORDS = ("1 курс", "системы и")

logger = logging.getLogger(__name__)

_DATE_RE = re.compile(r"(\d{2}\.\d{2}\.\d{4})")
_URL_RE = re.compile(r'https?://[^\s)>"\']+')
_TIME_RE = re.compile(r"(\d{1,2}:\d{2})\s*[-–—]\s*(\d{1,2}:\d{2})")
_WEEKDAY_RU = {
    0: "понедельник",
    1: "вторник",
    2: "среда",
    3: "четверг",
    4: "пятница",
    5: "суббота",
    6: "воскресенье",
}
_MONTH_RU = {
    1: "января",
    2: "февраля",
    3: "марта",
    4: "апреля",
    5: "мая",
    6: "июня",
    7: "июля",
    8: "августа",
    9: "сентября",
    10: "октября",
    11: "ноября",
    12: "декабря",
}


def _normalize_subgroup_name(subgroup: str | None) -> str | None:
    if not subgroup:
        return None
    subgroup_lower = subgroup.lower()
    if subgroup_lower.endswith("-1") or "25-1" in subgroup_lower:
        return "ПС-25-1"
    if subgroup_lower.endswith("-2") or "25-2" in subgroup_lower:
        return "ПС-25-2"
    return subgroup


@dataclass
class Lesson:
    subgroup: str
    date_value: date
    start_time: str
    end_time: str
    title: str
    link: str | None


@dataclass
class ScheduleFileRef:
    name: str
    url: str
    date_value: date | None


def _normalize_url(href: str) -> str:
    if href.startswith("//"):
        return f"https:{href}"
    if href.startswith("/"):
        return f"https://www.hse.ru{href}"
    return href


def _extract_date(raw: str) -> date | None:
    match = _DATE_RE.search(raw)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%d.%m.%Y").date()
    except ValueError:
        return None


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    return str(value).strip()


def _excel_serial_to_date(value: float) -> date | None:
    try:
        if value < 1:
            return None
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=float(value))).date()
    except Exception:
        return None


def _excel_fraction_to_time(value: float) -> str | None:
    try:
        if value < 0 or value >= 1:
            return None
        total_minutes = int(round(value * 24 * 60))
        hours = (total_minutes // 60) % 24
        minutes = total_minutes % 60
        return f"{hours:02d}:{minutes:02d}"
    except Exception:
        return None


def _parse_row_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (float, int)):
        serial_date = _excel_serial_to_date(float(value))
        if serial_date:
            return serial_date
    text = _cell_to_text(value)
    if not text:
        return None

    date_match = _DATE_RE.search(text)
    if date_match:
        try:
            return datetime.strptime(date_match.group(1), "%d.%m.%Y").date()
        except ValueError:
            pass

    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_time_range(value: object) -> tuple[str, str] | None:
    if isinstance(value, (float, int)):
        one_time = _excel_fraction_to_time(float(value))
        if one_time:
            return one_time, one_time

    text = _cell_to_text(value)
    if not text:
        return None
    match = _TIME_RE.search(text)
    if not match:
        return None
    return match.group(1), match.group(2)


def _cleanup_lesson_title(raw: str) -> str:
    title = re.sub(r"\s+", " ", raw).strip(" -\n\t")
    title = re.sub(r"^1\s*курс\.?\s*", "", title, flags=re.IGNORECASE)
    return title.strip()


def _parse_cell_entries(value: object) -> list[tuple[str, str | None]]:
    text = _cell_to_text(value)
    if not text:
        return []

    matches = list(_URL_RE.finditer(text))
    if not matches:
        title = _cleanup_lesson_title(text)
        return [(title, None)] if title else []

    result: list[tuple[str, str | None]] = []
    start = 0
    for match in matches:
        title_raw = text[start : match.start()]
        title = _cleanup_lesson_title(title_raw)
        link = match.group(0)
        if title or link:
            result.append((title or "Занятие", link))
        start = match.end()

    tail = _cleanup_lesson_title(text[start:])
    if tail:
        result.append((tail, None))

    return result


def _load_workbook_with_fallback(file_bytes: bytes) -> Any:
    try:
        return load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except Exception as openpyxl_error:
        logger.warning("openpyxl не смог открыть файл, пробую xlrd: %s", openpyxl_error)
        try:
            return xlrd.open_workbook(file_contents=file_bytes)
        except Exception as xlrd_error:
            raise ValueError(
                "Не удалось открыть файл расписания как XLSX/XLS"
            ) from xlrd_error


def _iter_sheet_titles(workbook: Any) -> list[str]:
    if hasattr(workbook, "worksheets"):
        return [sheet.title for sheet in workbook.worksheets]
    return workbook.sheet_names()


def _iter_sheet_rows(
    workbook: Any, sheet_title: str
) -> list[tuple[object, object, object, object]]:
    if hasattr(workbook, "worksheets"):
        sheet = workbook[sheet_title]
        result: list[tuple[object, object, object, object]] = []
        for row in range(4, sheet.max_row + 1):
            result.append(
                (
                    sheet.cell(row=row, column=1).value,
                    sheet.cell(row=row, column=2).value,
                    sheet.cell(row=row, column=3).value,
                    sheet.cell(row=row, column=4).value,
                )
            )
        return result

    sheet = workbook.sheet_by_name(sheet_title)
    result = []
    for row in range(3, sheet.nrows):
        row_values = sheet.row_values(row)
        c1 = row_values[0] if len(row_values) > 0 else None
        c2 = row_values[1] if len(row_values) > 1 else None
        c3 = row_values[2] if len(row_values) > 2 else None
        c4 = row_values[3] if len(row_values) > 3 else None
        result.append((c1, c2, c3, c4))
    return result


async def fetch_timetable_page() -> str:
    async with httpx.AsyncClient(follow_redirects=True, timeout=30.0) as client:
        response = await client.get(TIMETABLE_URL)
        response.raise_for_status()
        return response.text


def find_actual_schedule_file(page_html: str) -> ScheduleFileRef:
    soup = BeautifulSoup(page_html, "lxml")
    candidates: list[ScheduleFileRef] = []
    for a in soup.find_all("a", href=True):
        href = (a.get("href") or "").strip()
        if not href:
            continue
        text = (a.get_text(" ", strip=True) or "").strip()
        if not text:
            continue

        lower_text = text.lower()
        if not all(keyword in lower_text for keyword in FILE_KEYWORDS):
            continue

        lower_href = href.lower()
        data_hse_file = (a.get("data-hse-file") or "").upper()
        if (
            data_hse_file != "XLS"
            and ".xls" not in lower_href
            and ".xlsx" not in lower_href
        ):
            continue

        normalized_url = _normalize_url(href)
        file_date = _extract_date(text) or _extract_date(normalized_url)
        candidates.append(
            ScheduleFileRef(name=text, url=normalized_url, date_value=file_date)
        )

    if not candidates:
        raise ValueError("Не найден файл расписания по ключам: '1 курс' и 'системы и'")

    candidates.sort(key=lambda item: item.date_value or date.min, reverse=True)
    chosen = candidates[0]
    logger.info("Выбран файл расписания: %s (%s)", chosen.name, chosen.url)
    return chosen


async def download_schedule_file(file_url: str) -> bytes:
    async with httpx.AsyncClient(follow_redirects=True, timeout=30.0) as client:
        response = await client.get(file_url)
        response.raise_for_status()
        return response.content


def parse_lessons_for_today(
    file_bytes: bytes,
    today: date | None = None,
    subgroup: str | None = None,
) -> list[Lesson]:
    current_day = today or date.today()
    workbook = _load_workbook_with_fallback(file_bytes)
    lessons: list[Lesson] = []

    parsed_sheets = 0
    for sheet_title in _iter_sheet_titles(workbook):
        if "неделя" not in sheet_title.lower():
            continue
        parsed_sheets += 1

        current_row_date: date | None = None
        for cell_a, cell_b, cell_c, cell_d in _iter_sheet_rows(workbook, sheet_title):
            row_date = _parse_row_date(cell_a)
            if row_date:
                current_row_date = row_date

            if current_row_date != current_day:
                continue

            if current_row_date is None:
                continue

            time_range = _parse_time_range(cell_b)
            if not time_range:
                continue

            start_time, end_time = time_range
            entries_c = _parse_cell_entries(cell_c)
            entries_d = _parse_cell_entries(cell_d)

            if entries_c and not entries_d:
                subgroup_entries = ((entries_c, "ПС-25-1"), (entries_c, "ПС-25-2"))
            elif entries_d and not entries_c:
                subgroup_entries = ((entries_d, "ПС-25-1"), (entries_d, "ПС-25-2"))
            else:
                subgroup_entries = ((entries_c, "ПС-25-1"), (entries_d, "ПС-25-2"))

            for entries, subgroup in subgroup_entries:
                for title, link in entries:
                    if not title and not link:
                        continue
                    lessons.append(
                        Lesson(
                            subgroup=subgroup,
                            date_value=current_row_date,
                            start_time=start_time,
                            end_time=end_time,
                            title=title or "Занятие",
                            link=link,
                        )
                    )

    subgroup_normalized = _normalize_subgroup_name(subgroup)
    if subgroup_normalized:
        lessons = [item for item in lessons if item.subgroup == subgroup_normalized]

    lessons.sort(key=lambda item: item.start_time)
    logger.info(
        "Найдено занятий на %s: %s (листы: %s)",
        current_day,
        len(lessons),
        parsed_sheets,
    )
    return lessons


def format_today_schedule_message(
    lessons: list[Lesson],
    target_date: date | None = None,
    subgroup: str | None = None,
) -> str:
    current_day = target_date or date.today()
    if not lessons:
        return ""

    date_human = f"{current_day.day} {_MONTH_RU[current_day.month]}"
    weekday = _WEEKDAY_RU[current_day.weekday()]
    subgroup_normalized = _normalize_subgroup_name(subgroup)
    group_part = f" для группы {subgroup_normalized}" if subgroup_normalized else ""
    lines = [
        f"Ждём вас сегодня, {date_human} ({weekday}), на занятиях{group_part}📚",
        "",
    ]

    for item in lessons:
        title = html.escape(item.title)
        lines.append(f"⏰ {item.start_time}-{item.end_time} | {title}")
        if item.link:
            safe_link = html.escape(item.link)
            lines.append(f' └ <a href="{safe_link}">Ссылка на подключение</a>')
        lines.append("")

    return "\n".join(lines).strip()


async def get_today_schedule_message(
    today: date | None = None,
    subgroup: str | None = None,
) -> str:
    page_html = await fetch_timetable_page()
    file_ref = find_actual_schedule_file(page_html)
    file_bytes = await download_schedule_file(file_ref.url)
    target_date = today or date.today()
    lessons = parse_lessons_for_today(file_bytes, target_date, subgroup=subgroup)
    return format_today_schedule_message(lessons, target_date, subgroup=subgroup)
